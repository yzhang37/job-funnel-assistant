import json
from datetime import datetime
from pathlib import Path

import openpyxl


SOURCE_XLSX = Path("/Users/l/Downloads/工作管理系统.xlsx")
OUTPUT_JSON = Path("/Users/l/Projects/找工作/data/processed/wolai_import_batches.json")

STAGE_URLS = {
    "Pending Referal": "https://www.notion.so/3488df94deab81228578f5140fa4cabe",
    "Pending Referral": "https://www.notion.so/3488df94deab81228578f5140fa4cabe",
    "Applied": "https://www.notion.so/3488df94deab815eb3fef8e344a5ffef",
    "Recruiter Contacted": "https://www.notion.so/3488df94deab8148a159c7c53c5d176d",
    "Screening": "https://www.notion.so/3488df94deab816ea568c91757936acd",
    "Interviewing": "https://www.notion.so/3488df94deab8163b10acc5d33b85f4b",
    "Offer": "https://www.notion.so/3488df94deab81fc930dcbd1fa3f522c",
    "Ghost": "https://www.notion.so/3488df94deab8136aa9cfd485a47f2be",
    "Rejected": "https://www.notion.so/3488df94deab81f1b504f3527a23ca3e",
}


def parse_excel_date(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return {
            "start": value.isoformat(timespec="minutes"),
            "is_datetime": 1,
        }
    text = str(value).strip()
    if not text:
        return None
    formats = [
        ("%Y/%m/%d %H:%M", 1),
        ("%Y-%m-%d %H:%M", 1),
        ("%Y/%m/%d", 0),
        ("%Y-%m-%d", 0),
    ]
    for fmt, is_datetime in formats:
        try:
            parsed = datetime.strptime(text, fmt)
            if is_datetime:
                return {
                    "start": parsed.isoformat(timespec="minutes"),
                    "is_datetime": 1,
                }
            return {
                "start": parsed.date().isoformat(),
                "is_datetime": 0,
            }
        except ValueError:
            continue
    raise ValueError(f"Unsupported date format: {value!r}")


def build_page(row, headers):
    data = dict(zip(headers, row))
    props = {
        "Title": data["Title"],
        "Company": data["Company"],
        "External ID": data["External ID"],
        "Source": data["Source"],
        "Canonical URL": data["Canonical URL"],
        "POC / Recruiter": data["POC / Recruiter"],
        "Notes": data["Notes"],
        "Country": data["Country"],
    }

    stage = data["Stage"]
    if stage:
        stage_url = STAGE_URLS.get(stage)
        if not stage_url:
            raise KeyError(f"Unknown stage: {stage}")
        props["Stage"] = json.dumps([stage_url], ensure_ascii=False)

    applied_at = parse_excel_date(data["Applied At"])
    if applied_at:
        props["date:Applied At:start"] = applied_at["start"]
        props["date:Applied At:is_datetime"] = applied_at["is_datetime"]

    last_activity = parse_excel_date(data["Last Activity"])
    if last_activity:
        props["date:Last Activity:start"] = last_activity["start"]
        props["date:Last Activity:is_datetime"] = last_activity["is_datetime"]

    snooze_until = parse_excel_date(data["Snooze Until"])
    if snooze_until:
        props["date:Snooze Until:start"] = snooze_until["start"]
        props["date:Snooze Until:is_datetime"] = snooze_until["is_datetime"]

    original_create = parse_excel_date(data["Create Time"])
    if original_create:
        props["date:Original Create Time:start"] = original_create["start"]
        props["date:Original Create Time:is_datetime"] = original_create["is_datetime"]

    return {"properties": {k: v for k, v in props.items() if v not in (None, "")}}


def chunked(items, size):
    for start in range(0, len(items), size):
        yield items[start : start + size]


def main():
    wb = openpyxl.load_workbook(SOURCE_XLSX, data_only=False)
    ws = wb[wb.sheetnames[0]]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = [tuple(cell.value for cell in row) for row in ws.iter_rows(min_row=2)]

    pages = [build_page(row, headers) for row in rows if row[0]]
    batches = list(chunked(pages, 50))

    OUTPUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_JSON.write_text(
        json.dumps(
            {
                "count": len(pages),
                "batch_count": len(batches),
                "batches": batches,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    preview = {
        "count": len(pages),
        "batch_count": len(batches),
        "first_page": pages[0],
        "last_page": pages[-1],
    }
    print(json.dumps(preview, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
